import openpyxl

workbbok=openpyxl.load_workbook("D:\Python\Workspace\8thNov_pytestFramework\TestData\SwagLab.xlsx")
sheet=workbbok['Sheet1']

#get row & col size
print(sheet.max_row)
print(sheet.max_column)


#get specific data
print(sheet["A1"].value)
print(sheet["B2"].value)

print("---")

print(sheet.cell(1,1).value)
print(sheet.cell(2,2).value)


#get data & convert into int/float
dataInInt=int(sheet.cell(1,2).value)
print(dataInInt+5)


dataInfloat=float(sheet.cell(1,4).value)
print(dataInfloat)

sheet1=workbbok["Sheet3"]
print(sheet1.cell(3,1).value)
print(sheet1.cell(3,1).value)
print(sheet1.cell(3,1).value)
print(sheet1.cell(3,1).value)

# firstProductPrice = "$29.99"
# firstProductPrice=firstProductPrice[1:]
# print(firstProductPrice)


