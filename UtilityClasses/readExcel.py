import openpyxl
class ReadExcel:

   @staticmethod
   def getExcelData(rowIndex,colIndex):
      workbook = openpyxl.load_workbook(
          "D:\Python\Workspace\FrameworkSetup_Pytest\TestData\8thNovSwagLab.xlsx")
      sheet = workbook['SwagLab']

      data = sheet.cell(rowIndex, colIndex).value
      return data

   @staticmethod
   def getExcelDataInt(rowIndex, colIndex):
      workbook = openpyxl.load_workbook(
         "D:\\Other_Notes\\Python\\Workspace\\1stFebPythonSelenium_Framework\\TestData\\Data.xlsx")
      sheet = workbook['Sheet2']

      s2 = sheet.cell(rowIndex, colIndex).value
      s2 = int(s2)  # convert String to int
      return s2
